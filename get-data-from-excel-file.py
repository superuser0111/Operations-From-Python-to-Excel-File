# Import the openpyxl library
import openpyxl

# Load the Excel file to create an ExcelFile object
excel_File = openpyxl.load_workbook("gd-names-surnames.xlsx")

# Select a specific page in the Excel file (default name assumed to be "Sheet1")
page = excel_File["Sheet1"]

# Create an empty list to store names and surnames
names_Surnames = []

# Get the indices of the last row and column in the page
max_Row = page.max_row
max_Column = page.max_column

# Iterate through the data table, processing each row
for row in range(2, max_Row + 1):

    # Create an empty sub-list for each row
    row_Data = []

    # Iterate through each column in the row
    for column in range(1, max_Column + 1):

        # Retrieve the value in the cell and append it to the sub-list row_Data
        row_Data.append(page.cell(row, column).value)

    # Append the sub-list containing name and surname to the names_Surnames list
    names_Surnames.append(row_Data)

# Print the names and surnames to the console
for i in names_Surnames:

    print(f"Name: {i[0]} | Surname: {i[1]}")