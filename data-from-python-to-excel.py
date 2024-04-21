import openpyxl

# Load the Excel file
excel_File = openpyxl.load_workbook("dt-names-surnames.xlsx")

# Select the specific sheet in the Excel file
page = excel_File["Sheet1"]

# Define column names and data
column_Names = ["Names", "Surnames"]
data = [["Name1", "Surname1"], ["Name2", "Surname2"], ["Name3", "Surname3"], ["Name4", "Surname4"], ["Name5", "Surname5"]]

# Write column names to the first row
for i in column_Names:
    page.cell(row=1, column=column_Names.index(i) + 1, value=i)

# Write data to the sheet
for i in data:
    for j in i:
        page.cell(row=data.index(i) + 2, column=i.index(j) + 1, value=j)

# Save the changes to the Excel file
excel_File.save("dt-names-surnames.xlsx")